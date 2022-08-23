from aiogram import Bot, types
from aiogram.utils import executor
from aiogram.dispatcher import Dispatcher
from aiogram.types import ReplyKeyboardRemove, \
    ReplyKeyboardMarkup, KeyboardButton, \
    InlineKeyboardMarkup, InlineKeyboardButton
import openpyxl
import os
from datetime import datetime
import logging
from asyncio import get_event_loop
import time
import sys
logger = logging.getLogger(__name__)
logging.basicConfig(filename="log.log", level=logging.ERROR,format="%(asctime)s - %(levelname)s - %(name)s - %(message)s")
logger.error("Starting bot")

bot = Bot(token='5479649704:AAGuvYHgp_DvzVxV6uV8j4RnKwG1nMsa5rw')
dp = Dispatcher(bot=bot, loop=get_event_loop())

async def raed_excel():

        path = os.path.abspath(os.path.curdir)
        if not os.path.exists(f'{path}/documents'):
                os.makedirs(f'{path}/documents')

        path = os.path.abspath(f'{path}/documents')

        dir_list = [os.path.join(path, x) for x in os.listdir(path)]

        if len(dir_list) >= 1:
                # Создадим список из путей к файлам и дат их создания.
                date_list = [[x, os.path.getctime(x)] for x in dir_list]

                # Отсортируем список по дате создания в обратном порядке
                sort_date_list = sorted(date_list, key=lambda x: x[1], reverse=True)

                # Выведем первый элемент списка. Он и будет самым последним по дате
                list_excel = sort_date_list[0][0]


        file_to_read = openpyxl.load_workbook(list_excel, data_only=True)
        sheet = file_to_read['Таблица']

        s = []
        for row in range(3, sheet.max_row + 1):
                if 'None' not in str(sheet[row][10].value) and '\n' not in sheet[row][10].value:
                        s.append(sheet[row][10].value)

        from collections import Counter
        nunber = dict(Counter(s))
        return nunber

async def number_of_chats():
        path = os.path.abspath(os.path.curdir)
        if not os.path.exists(f'{path}/documents'):
                os.makedirs(f'{path}/documents')

        path = os.path.abspath(f'{path}/documents')

        dir_list = [os.path.join(path, x) for x in os.listdir(path)]

        if len(dir_list) >= 1:
                # Создадим список из путей к файлам и дат их создания.
                date_list = [[x, os.path.getctime(x)] for x in dir_list]

                # Отсортируем список по дате создания в обратном порядке
                sort_date_list = sorted(date_list, key=lambda x: x[1], reverse=True)

                # Выведем первый элемент списка. Он и будет самым последним по дате
                list_excel = sort_date_list[0][0]

        file_to_read = openpyxl.load_workbook(list_excel, data_only=True)
        sheet = file_to_read['Таблица']

        return sheet

async def data_time():
        path = os.path.abspath(os.path.curdir)
        if not os.path.exists(f'{path}/documents'):
                os.makedirs(f'{path}/documents')

        path = os.path.abspath(f'{path}/documents')

        dir_list = [os.path.join(path, x) for x in os.listdir(path)]

        if len(dir_list) >= 1:
                # Создадим список из путей к файлам и дат их создания.
                date_list = [[x, os.path.getctime(x)] for x in dir_list]

                # Отсортируем список по дате создания в обратном порядке
                sort_date_list = sorted(date_list, key=lambda x: x[1], reverse=True)

                # Выведем первый элемент списка. Он и будет самым последним по дате
                list_excel = sort_date_list[0][0]

        else:
                print('Файлов нет')

        ti_c = os.path.getctime(list_excel)
        ti_c = time.ctime(ti_c)
        return (ti_c)

@dp.message_handler(commands="start")
async def start(message: types.Message):
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        button_1 = types.KeyboardButton(text="Общая статистика")
        keyboard.add(button_1)
        button_2 = types.KeyboardButton(text="Своя")
        keyboard.add(button_2)
        button_2 = types.KeyboardButton(text="Без пометки ответ")
        keyboard.add(button_2)
        await message.answer("Главное меню", reply_markup=keyboard)

@dp.message_handler(lambda message: message.text == "Общая статистика")
async def General_statistics(message: types.Message):
        try:
                nunber = await raed_excel()
                s = []
                times=await data_time()
                await message.answer(f'Дата и время читаемого документа: \n{times}')
                for i in sorted(nunber.items(), key=lambda para: para[1], reverse=True):
                        s.append(f'{i[1]} - {i[0]}')
                if len(s)>0:
                        await message.answer('\n'.join(s))
                else:
                        await message.answer('Файл пуст')
        except UnboundLocalError:
                await message.answer('Файлов нет, загрузите файл Выгрузка.xlsx')
        except openpyxl.utils.exceptions.InvalidFileException:
                await message.answer('Файл не соответствует формату .xlsx,.xlsm,.xltx,.xltm')

@dp.message_handler(lambda message: message.text == "Своя")
async def Own(message: types.Message):
        try:
                nunber = await raed_excel()
                s = []
                for i in sorted(nunber.items(), key=lambda para: para[1], reverse=True):
                        if str(i[0]) != 'None' and '\n' not in i[0] :
                                s.append(f'{i[0]}')

                markup = InlineKeyboardMarkup()  # создаём клавиатуру
                markup.row_width = 1  # кол-во кнопок в строке
                for i in s:  # цикл для создания кнопок
                        markup.add(InlineKeyboardButton(i, callback_data=i))

                await bot.send_message(message.from_user.id, 'Выберите эксперта:', reply_markup=markup)
        except UnboundLocalError:
                await message.answer('Файлов нет, загрузите файл Выгрузка.xlsx')
        except openpyxl.utils.exceptions.InvalidFileException:
                await message.answer('Файл не соответствует формату .xlsx,.xlsm,.xltx,.xltm')

@dp.callback_query_handler(lambda call: True)
async def stoptopupcall(call: types.CallbackQuery):
        expert=call.data

        list_expert = await raed_excel()
        ####
        if expert in list_expert:
                namber_chat=(list_expert[expert])

        ####
        exel= await number_of_chats()
        ssilki = []
        ssilki2 = []
        chats_exit_time=0
        for row in range(3, exel.max_row + 1):
                p_key = str(exel[row][16].value)  # импортируемая ячейка, где row строка, а [x] столбец
                if str(exel[row][10].value) !='None' :
                        if len(str(p_key)) == 14:
                                dt = datetime.strptime(p_key, '%H:%M:%S.%f')
                        elif len(str(p_key))==7:
                                dt = datetime.strptime(p_key, '%H:%M:%S')

                        dt= dt.strftime("%M:%S")
                        if dt >='07:00' and exel[row][10].value == expert :
                                chats_exit_time+=1
                                ssilki.append(exel[row][1].value)

                        if expert in exel[row][10].value and '\n' in exel[row][10].value :
                                ssilki2.append(exel[row][1].value)

        times = await data_time()
        await bot.send_message(chat_id=call.from_user.id,text=f'Эксперт: {expert}\n'
                                                              f'Дата и время читаемого документа: \n{times}\n'
                                                              f'Количество обработанных чатов: {namber_chat}\n'
                                                              f'Количество чатов с превышением по времени: {chats_exit_time}\n'
                                                              f'Ссылки просроченных чатов: {ssilki}\n'
                                                              f'Ссылки на чаты с более чем одной пометкой: {ssilki2}')

@dp.message_handler(lambda message: message.text == "Без пометки ответ")
async def Unmarked_answer(message: types.Message):
        try:
                excel=await number_of_chats()
                s=[]
                for row in range(3, excel.max_row + 1):
                        p_key = str(excel[row][10].value)
                        if p_key == "None":
                                s.append(excel[row][1].value)
                await message.answer(f'Ссылки на чаты: {s}')
        except UnboundLocalError:
                await message.answer('Файлов нет, загрузите файл Выгрузка.xlsx')
        except openpyxl.utils.exceptions.InvalidFileException:
                await message.answer('Файл не соответствует формату .xlsx,.xlsm,.xltx,.xltm')

@dp.message_handler(content_types=types.ContentType.DOCUMENT)
async def scan_message(message: types.Message):

        destination = os.path.abspath(os.path.curdir)
        await message.document.download(destination)
        await message.answer("Файл загружен")

@dp.message_handler(commands="admin")
async def start_admin(message: types.Message):
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        button_1 = types.KeyboardButton(text="Рестарт")
        keyboard.add(button_1)
        button_2 = types.KeyboardButton(text="Количество файлов")
        keyboard.add(button_2)
        button_3 = types.KeyboardButton(text="Удалить каталог")
        keyboard.add(button_3)
        button_5 = types.KeyboardButton(text="Получить лог")
        keyboard.add(button_5)
        button_4 = types.KeyboardButton(text="Назад")
        keyboard.add(button_4)
        await message.answer("Режим Администратора", reply_markup=keyboard)

@dp.message_handler(lambda message: message.text == "Рестарт")
async def restarting(message: types.Message):
        await message.answer('Бот перезапущен')
        os.execl(sys.executable, sys.executable, *sys.argv)

@dp.message_handler(lambda message: message.text == "Количество файлов")
async def num_file(message: types.Message):
        path = os.path.abspath(os.path.curdir)
        if not os.path.exists(f'{path}/documents'):
                os.makedirs(f'{path}/documents')
        path = os.path.abspath(f'{path}/documents')
        dir_list = [os.path.join(path, x) for x in os.listdir(path)]
        await message.answer(f'Количество файлов в каталоге : {len(dir_list)}')

@dp.message_handler(lambda message: message.text == "Удалить каталог")
async def num_file(message: types.Message):
        import shutil
        path = os.path.abspath(os.path.curdir)
        if not os.path.exists(f'{path}/documents'):
                os.makedirs(f'{path}/documents')
        path = os.path.abspath(f'{path}/documents')
        shutil.rmtree(path)
        await message.answer('Каталог удален')

@dp.message_handler(lambda message: message.text == 'Получить лог')
async def logs(message: types.Message):
    await message.answer_document(open("log.log", "rb"))

@dp.message_handler(lambda message: message.text == 'Назад')
async def func_ex(message: types.Message):
    await start(message)
executor.start_polling(dp, skip_updates=True)